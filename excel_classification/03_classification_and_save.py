import pandas as pd
from openpyxl.reader.excel import load_workbook
from datetime import datetime
from openpyxl.styles import Font, Alignment

pd.set_option('display.max_columns', None)
import openpyxl


class ClassificationExcel:
    path = ''

    def __init__(self, order_xlsx_filename, partner_info_xlsx_file_name, path='result'):
        # 주문 목록
        df = pd.read_excel(order_xlsx_filename)
        df = df.rename(columns=df.iloc[1])  # 1번을 열제목으로 쓰겠다 라는 뜻
        '''
        0                                      NaN        NaN  ...         NaN         NaN
        1                                     주문번호       상품번호  ...          주소     주문시요구사항
        '''
        df = df.drop([df.index[0], df.index[1]])
        df = df.reset_index(drop=True)
        self.order_list = df
        self.path = path

        # 파트너 목록
        df_partners = pd.read_excel(partner_info_xlsx_file_name)

        self.brands = df_partners['브랜드'].to_list()
        self.partners = df_partners['업체명'].to_list()

    def classify(self):
        for i, row in self.order_list.iterrows():  # iterrows()는 행들을 반복할것이다
            brand_name = ''
            partner_name = ''
            for j in range(len(self.brands)):
                if self.brands[j] in row['상품명']:
                    brand_name = self.brands[j]
                    partner_name = self.partners[j]
                    break
            # print(f'{row["상품명"]}은 {brand_name} 브랜드 입니다. {idx_partners}번째')
            # print(f'업체명 : {partner_name}')
            if partner_name != '':
                df_filtered = self.order_list[self.order_list['상품명'].str.contains(brand_name)]
                df_filtered.to_excel(f'{self.path}/[패스트몰]{partner_name}.xlsx')
            else:
                print('없는 brand name', brand_name, row['상품명'])

    def set_count(self):
        file_name = 'result/[패스트몰]그레이스코퍼레이션.xlsx'
        wb = load_workbook(file_name)
        ws = wb.active
        print('value:', ws['A1'].value)

        # 개수 세기 max_row하면 head도 포함되어 -1 진행
        row_cnt = ws.max_row - 1
        print('cnt : ', row_cnt)

        # 열 삽입 2줄 추가
        ws.insert_rows(1)
        ws.insert_rows(1)

        now_day = datetime.now().strftime('%Y-%m-%d')
        # A1
        ws['A1'] = f'발송요청내역 [총{row_cnt}건]{now_day}'
        ws['A1'].font = Font(size=11, bold=True)
        ws.merge_cells('A1:U1')
        ws['A1'].alignment = Alignment(horizontal='left')

        wb.save(file_name)


if __name__ == '__main__':
    order_excel_filename = '주문목록20221112.xlsx'
    partner_info_excel_filename = '파트너목록.xlsx'
    ce = ClassificationExcel(order_excel_filename, partner_info_excel_filename)
    # ce.classify()
    ce.set_count()
