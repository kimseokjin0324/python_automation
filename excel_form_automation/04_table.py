from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook


# 엑셀파일 생성하는 코드
class WeeklyWorkPlan:
    wb = None  # wb 은 엑셀파일에 해당
    ws = None  # ws(워크시트)는 시트에 해당
    start_date = "2023-03-01"
    manager = "기본매니저"
    date_list = []
    days_of_week = []

    def __init__(self, start_date, manager, sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager

        # 날짜 생성
        self.set_date()
        self.set_title()
        self.set_table()

    def save(self, fileName):
        self.wb.save(fileName)
        print("엑셀 파일 생성 완료")

    def set_title(self):
        ws = self.ws
        ws['B2'] = '담당자'
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        # 제목
        ws['B5'] = '주간업무계획표'
        start_date = self.date_list[0]
        end_date = self.date_list[-1]
        ws['B6'] = f'({start_date} ~ {end_date})'

        # 셀병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

    def set_table(self):
        ws = self.ws
        ws['B8'] = '날짜'
        col_names = ['날짜', '요일', '시간', '일정', '비고']

        # column 명
        for i in range(len(col_names)):
            ws.cell(row=8, column=i + 2).value = col_names[i]

        # 날짜및 요일 뿌리기
        for i in range(len(self.date_list)):
            ws.cell(row=9 + i, column=2).value = self.date_list[i]
            ws.cell(row=9 + i, column=3).value = self.days_of_week[i]

        # 행삽입하기
        ws.insert_rows(10,4)
        ws.insert_rows(15,4)
        ws.insert_rows(20,4)
        ws.insert_rows(25,4)
        ws.insert_rows(30,4)


    def set_date(self, days=6):
        # start_date +6일 해 줘야함 -> end_date 구하기
        end_date = datetime.strptime(self.start_date, "%Y-%m-%d") + timedelta(days=days)

        # 6일치 날짜 리스트만들기
        week = pd.date_range(start=self.start_date, end=end_date.strftime("%Y-%m-%d"))
        self.date_list = week.strftime("%Y-%m-%d").to_list()
        self.days_of_week = week.strftime("%A").to_list()
        print('end_date ', end_date)
        print('week', week)
        print('date_list', self.date_list)
        print("days_of_week", self.days_of_week)


if __name__ == '__main__':
    wwp = WeeklyWorkPlan('2023-03-08', '김석진')
    wwp.save('주간업무계획표.xlsx')
