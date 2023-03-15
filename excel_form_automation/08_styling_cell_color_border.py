from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# 엑셀파일 생성하는 코드
class WeeklyWorkPlan:
    wb = None  # wb 은 엑셀파일에 해당
    ws = None  # ws(워크시트)는 시트에 해당
    start_date = "2023-03-01"
    manager = "기본매니저"
    date_list = []
    days_of_week = []

    def __init__(self, start_date, manager, days=5, sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager

        # 날짜 생성
        self.set_date(days=days)
        self.set_title()
        self.set_table()
        self.set_style()

    def save(self, fileName):
        self.wb.save(fileName)
        print("엑셀 파일 생성 완료")

    def set_title(self):
        ws = self.ws
        ws['B2'] = '담당자'
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        # 제목
        ws['B5'] = '주간업무계획표'
        start_date = self.date_list[0]
        end_date = self.date_list[-1]
        ws['B6'] = f'({start_date} ~ {end_date})'

        # 셀병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

    def set_table(self):
        ws = self.ws
        ws['B8'] = '날짜'
        col_names = ['날짜', '요일', '시간', '일정', '비고']

        # column 명
        for i in range(len(col_names)):
            ws.cell(row=8, column=i + 2).value = col_names[i]

        # 날짜및 요일 뿌리기
        for i in range(len(self.date_list)):
            ws.cell(row=9 + (i * 5), column=2).value = self.date_list[i]
            ws.cell(row=9 + (i * 5), column=3).value = self.days_of_week[i]
            ws.merge_cells(f'B{9 + i * 5}:B{13 + i * 5}')  # 날짜
            ws.merge_cells(f'C{9 + i * 5}:C{13 + i * 5}')  # 요일
            ws.merge_cells(f'F{9 + i * 5}:F{13 + i * 5}')  # 비고

    def set_style(self):
        ws = self.ws

        # 너비 설정하기 (column_dimensions)
        # A열 너비
        ws.column_dimensions['A'].width = 5

        # B C D E F열 너비 열제목
        for i in range(2, 7):
            ws.column_dimensions[get_column_letter(i)].width = 15
            ws[f'{get_column_letter(i)}8'].font = Font(name='맑은 고딕', bold=True)
            ws[f'{get_column_letter(i)}8'].alignment = Alignment(horizontal='center', vertical='center')
            # fill 색상 fgColor에 색상코드를 넣으면 됨
            ws[f'{get_column_letter(i)}8'].fill = PatternFill(fgColor='E2EFDA', fill_type='solid')

        # E열 너비
        ws.column_dimensions['E'].width = 40

        # 제목 글꼴,사이즈
        ws['B5'].font = Font(name='맑은 고딕', size=28, bold=True)

        # 가운데 정렬
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')

        # 날짜요일 가운데 정렬
        for i in range(9, 40, 5):
            ws[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{i}'].alignment = Alignment(horizontal='center', vertical='center')

        # 담당자,시작일 색칠
        ws['B2'].fill = PatternFill(fgColor='E2EFDA', fill_type='solid')
        ws['B3'].fill = PatternFill(fgColor='E2EFDA', fill_type='solid')

        # 테두리 설정 Border
        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin')
                              )
        ws['B2'].border = border_style
        ws['C2'].border = border_style
        ws['B3'].border = border_style
        ws['C3'].border = border_style

        # 표영역 Border iter_cols(컬럼을 반복)

        for col in ws.iter_cols(min_row=8, min_col=2, max_row=len(self.date_list) * 5 +8, max_col=6):
         for cell in col:
            cell.border = border_style

    def set_date(self, days=6):
        # start_date +6일 해 줘야함 -> end_date 구하기
        end_date = datetime.strptime(self.start_date, "%Y-%m-%d") + timedelta(days=days)

        # 6일치 날짜 리스트만들기
        week = pd.date_range(start=self.start_date, end=end_date.strftime("%Y-%m-%d"))
        self.date_list = week.strftime("%Y-%m-%d").to_list()
        self.days_of_week = week.strftime("%A").to_list()
        print('end_date ', end_date)
        print('week', week)
        print('date_list', self.date_list)
        print("days_of_week", self.days_of_week)


if __name__ == '__main__':
    wwp = WeeklyWorkPlan('2023-03-08', '김석진', days=5)
    wwp.save('주간업무계획표.xlsx')
