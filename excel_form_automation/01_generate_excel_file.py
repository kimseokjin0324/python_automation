from openpyxl import Workbook


# 엑셀파일 생성하는 코드
class WeeklyWorkPlan:
    wb = None  # wb 은 엑셀파일에 해당
    ws = None  # ws(워크시트)는 시트에 해당
    start_date = "2023-03-01"
    manager = "기본매니저"

    def __init__(self, start_date, manager, sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager
        self.set_title()

    def save(self, fileName):
        self.wb.save(fileName)
        print("엑셀 파일 생성 완료")

    def set_title(self):
        self.ws['B2'] = '담당자'
        self.ws['C2'] = self.manager
        self.ws['B3'] = '시작일'
        self.ws['C3'] = self.start_date


if __name__ == '__main__':
    wwp = WeeklyWorkPlan('2023-03-08','김석진')
    wwp.save('주간업무계획표.xlsx')
