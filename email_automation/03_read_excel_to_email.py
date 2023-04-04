import os
import smtplib  # 파이썬 내장라이브러리
from email.mime.text import MIMEText

from openpyxl.reader.excel import load_workbook


class EmailSender:
    email_addr = None
    password = None
    smtp_server_map = {
        'gmail.com': 'smtp.gmail.com',
        'naver.com': 'smtp.naver.com'
    }
    smtp_server = None

    def __init__(self, email_addr, password):
        print('생성자')
        self.email_addr = email_addr
        self.password = password
        self.smtp_server = self.smtp_server_map[email_addr.split('@')[1]]

    def send_email(self, msg, from_addr, to_addr, subject):
        """
        :param msg: 보낼 메세지
        :param from_addr: 보내는 곳 사람
        :param to_addr: 받는 사람
        :return: 없을 예정
        """
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            smtp.starttls()
            smtp.login(self.email_addr, self.password)
            # 네이버는 구글과다르게 다른 작업이 필요하다
            msg = MIMEText(msg)
            msg['From'] = from_addr
            msg['To'] = to_addr
            msg['Subject'] = subject
            # 로그인 한 이후 이메일 보내기
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print(f'to_addr:{to_addr}로 이메일 전송이 완료되었습니다.')

    def send_all_emails(self, filename):
        print(f'{filename}에 있는 이메일과 내용을 이용해 메일을 보내기')
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_row =2):
            if row[0].value !=None:
                print(row[0].value,row[1].value,row[2].value)
                self.send_email(row[2].value, from_addr=self.email_addr, to_addr=row[0].value,
                      subject=row[1].value)


if __name__ == '__main__':
    # es = EmailSender('kimseokjin0324@gmail.com',os.getenv('MY_GMAIL_PASSWORD'))
    es = EmailSender('sksmstjrwls1@naver.com', os.getenv('MY_NAVER_PASSWORD'))
    # es.send_email(' 테스트 입니다 \n 네이버에서 보냄.', from_addr='sksmstjrwls1@naver.com', to_addr='kimseokjin0324@gmail.com',
    #               subject='이메일 전송 테스트 진행3')

    es.send_all_emails('이메일 리스트.xlsx')
