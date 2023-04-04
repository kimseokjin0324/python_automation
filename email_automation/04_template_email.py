import os
import smtplib  # 파이썬 내장라이브러리
from email.mime.text import MIMEText
from email.utils import formataddr

from openpyxl.reader.excel import load_workbook


class EmailSender:
    email_addr = None
    password = None
    manager_name =None
    smtp_server_map = {
        'gmail.com': 'smtp.gmail.com',
        'naver.com': 'smtp.naver.com'
    }
    smtp_server = None

    def __init__(self, email_addr, password,manager_name):
        print('생성자')
        self.email_addr = email_addr
        self.password = password
        self.manager_name =manager_name
        self.smtp_server = self.smtp_server_map[email_addr.split('@')[1]]

    def send_email(self, msg, from_addr, to_addr,receiver_name, subject):
        """
        :param msg: 보낼 메세지
        :param from_addr: 보내는 곳 사람
        :param to_addr: 받는 사람
        :return: 없을 예정
        """
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            smtp.starttls()
            smtp.login(self.email_addr, self.password)
            # 네이버는 구글과다르게 다른 작업이 필요하다
            msg = MIMEText(msg)
            msg['From'] = formataddr((self.manager_name,from_addr)) #메일에 보내는 이메일에 보내는사람 추가하기
            msg['To'] = formataddr((receiver_name,to_addr))
            msg['Subject'] = subject
            # 로그인 한 이후 이메일 보내기
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print(f'to_addr:{to_addr}로 이메일 전송이 완료되었습니다.')

    def send_all_emails(self, filename):
        print(f'{filename}에 있는 이메일과 내용을 이용해 메일을 보내기')
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_row =2):
            # 교체하는부분은 %교체할부분%으로 세팅해서 중복이 되지않게 템플릿화
            temp1 ="""
                안녕하세요 %받는분%님 패스트몰 %매니저_이름% 입니다.
                귀사에 무궁한 발전을 기원합니다.
                금일 쇼핑몰로 주문들어온 주문건들을 보내드립니다. 
                확인해보시고 발주 부탁드립니다.
                후지쯔코리아솔루션서비스 %매니저_이름% 드림
                 
            """
            if row[0].value !=None:
                print(row[0].value,row[1].value,row[2].value)
                temp1 =  temp1.replace('%받는분%',row[2].value)
                temp1 =  temp1.replace('%매니저_이름%',self.manager_name)
                self.send_email(msg = temp1,
                                from_addr=self.email_addr,
                                to_addr=row[0].value,
                                receiver_name=row[2].value,
                                subject=row[1].value)


if __name__ == '__main__':
    # es = EmailSender('kimseokjin0324@gmail.com',os.getenv('MY_GMAIL_PASSWORD'))
    es = EmailSender('sksmstjrwls1@naver.com', os.getenv('MY_NAVER_PASSWORD'),manager_name= '김석진')
    # es.send_email(' 테스트 입니다 \n 네이버에서 보냄.', from_addr='sksmstjrwls1@naver.com', to_addr='kimseokjin0324@gmail.com',
    #               subject='이메일 전송 테스트 진행3')

    es.send_all_emails('이메일 리스트_with_name.xlsx')
